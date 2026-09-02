"""
Genera 5 plantillas Excel premium para Perú
- Dashboard Financiero
- Control de Ventas
- Planilla de Personal
- Control de Inventario
- Flujo de Caja
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, DataBarRule, IconSetRule
from openpyxl.worksheet.datavalidation import DataValidation
import os

# ── Colores corporativos ──
DARK_BG = "1A1A2E"
GREEN_PRIMARY = "00B894"
GREEN_DARK = "00897B"
YELLOW_ACCENT = "FDCB6E"
ORANGE_ACCENT = "E17055"
RED_ALERT = "D63031"
BLUE_INFO = "0984E3"
PURPLE = "6C5CE7"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F6FA"
GRAY_TEXT = "636E72"

# ── Estilos base ──
header_font = Font(name='Calibri', bold=True, size=12, color=WHITE)
title_font = Font(name='Calibri', bold=True, size=16, color=DARK_BG)
subtitle_font = Font(name='Calibri', bold=True, size=11, color=GRAY_TEXT)
data_font = Font(name='Calibri', size=10)
number_font = Font(name='Calibri', size=10, bold=True)

header_fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type='solid')
green_fill = PatternFill(start_color=GREEN_PRIMARY, end_color=GREEN_PRIMARY, fill_type='solid')
yellow_fill = PatternFill(start_color=YELLOW_ACCENT, end_color=YELLOW_ACCENT, fill_type='solid')
light_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type='solid')

thin_border = Border(
    left=Side(style='thin', color='DFE6E9'),
    right=Side(style='thin', color='DFE6E9'),
    top=Side(style='thin', color='DFE6E9'),
    bottom=Side(style='thin', color='DFE6E9')
)

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center')

PEN_FORMAT = '#,##0.00" S/"'
PERCENT_FORMAT = '0.0%'


def style_header_row(ws, row, max_col, fill=None):
    """Aplica estilo de encabezado a una fila"""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = fill or header_fill
        cell.alignment = center_align
        cell.border = thin_border


def style_data_area(ws, start_row, end_row, max_col):
    """Aplica estilo a área de datos"""
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = center_align
            if row % 2 == 0:
                cell.fill = light_fill


def add_title_banner(ws, title, subtitle, row=1):
    """Agrega un banner de título profesional"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(name='Calibri', bold=True, size=18, color=WHITE)
    cell.fill = PatternFill(start_color=GREEN_DARK, end_color=GREEN_DARK, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 45

    ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=8)
    cell2 = ws.cell(row=row+1, column=1, value=subtitle)
    cell2.font = Font(name='Calibri', size=11, color=WHITE, italic=True)
    cell2.fill = PatternFill(start_color=GREEN_PRIMARY, end_color=GREEN_PRIMARY, fill_type='solid')
    cell2.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row+1].height = 25


def add_kpi_card(ws, row, col, label, value, color):
    """Agrega una tarjeta KPI"""
    # Label
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    cell_label = ws.cell(row=row, column=col, value=label)
    cell_label.font = Font(name='Calibri', size=9, color=WHITE)
    cell_label.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    cell_label.alignment = center_align

    # Value
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
    cell_val = ws.cell(row=row+1, column=col, value=value)
    cell_val.font = Font(name='Calibri', bold=True, size=14, color=WHITE)
    cell_val.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    cell_val.alignment = center_align
    ws.row_dimensions[row+1].height = 30


# ══════════════════════════════════════════════════════════════
# 1. DASHBOARD FINANCIERO
# ══════════════════════════════════════════════════════════════
def crear_dashboard_financiero():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = GREEN_PRIMARY

    # Configurar columnas
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 16

    # Título
    add_title_banner(ws, "📊 DASHBOARD FINANCIERO", "Control total de tu negocio — Año 2025")

    # KPIs principales
    row = 4
    add_kpi_card(ws, row, 1, "INGRESOS TOTALES", 285750, GREEN_PRIMARY)
    add_kpi_card(ws, row, 3, "GASTOS TOTALES", 178420, ORANGE_ACCENT)
    add_kpi_card(ws, row, 5, "UTILIDAD NETA", 107330, BLUE_INFO)
    add_kpi_card(ws, row, 7, "MARGEN %", "37.6%", PURPLE)

    # Datos mensuales
    row = 8
    ws.cell(row=row, column=1, value="MES").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=2, value="INGRESOS").font = header_font
    ws.cell(row=row, column=2).fill = header_fill
    ws.cell(row=row, column=3, value="GASTOS").font = header_font
    ws.cell(row=row, column=3).fill = header_fill
    ws.cell(row=row, column=4, value="UTILIDAD").font = header_font
    ws.cell(row=row, column=4).fill = header_fill
    ws.cell(row=row, column=5, value="MARGEN %").font = header_font
    ws.cell(row=row, column=5).fill = header_fill
    ws.cell(row=row, column=6, value="IGV 18%").font = header_font
    ws.cell(row=row, column=6).fill = header_fill
    ws.cell(row=row, column=7, value="RENTA 1.5%").font = header_font
    ws.cell(row=row, column=7).fill = header_fill

    meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
             "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    ingresos = [18500, 22300, 25800, 24100, 27600, 26400, 29800, 31200, 28900, 25700, 23400, 22050]
    gastos = [12400, 14800, 16200, 15900, 17100, 16800, 18200, 19500, 17800, 16100, 14900, 14720]

    for i, mes in enumerate(meses):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=mes)
        ws.cell(row=r, column=2, value=ingresos[i])
        ws.cell(row=r, column=3, value=gastos[i])
        ws.cell(row=r, column=4).value = f"=B{r}-C{r}"
        ws.cell(row=r, column=5).value = f"=IF(B{r}>0,D{r}/B{r},0)"
        ws.cell(row=r, column=6).value = f"=B{r}*0.18"
        ws.cell(row=r, column=7).value = f"=B{r}*0.015"

        # Formato de moneda
        for col in [2, 3, 4, 6, 7]:
            ws.cell(row=r, column=col).number_format = '#,##0'
        ws.cell(row=r, column=5).number_format = '0.0%'

    style_data_area(ws, row+1, row+12, 7)

    # Totales
    total_row = row + 13
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, size=11)
    for col in [2, 3, 4, 6, 7]:
        ws.cell(row=total_row, column=col).value = f"=SUM({get_column_letter(col)}{row+1}:{get_column_letter(col)}{row+12})"
        ws.cell(row=total_row, column=col).number_format = '#,##0'
        ws.cell(row=total_row, column=col).font = Font(bold=True, size=11)
    ws.cell(row=total_row, column=5).value = f"=IF(B{total_row}>0,D{total_row}/B{total_row},0)"
    ws.cell(row=total_row, column=5).number_format = '0.0%'

    # Gráfico de barras - Ingresos vs Gastos
    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "Ingresos vs Gastos Mensuales"
    chart1.y_axis.title = "Soles (S/)"
    chart1.x_axis.title = "Mes"
    chart1.style = 10
    chart1.width = 20
    chart1.height = 12

    data = Reference(ws, min_col=2, min_row=row, max_col=3, max_row=row+12)
    cats = Reference(ws, min_col=1, min_row=row+1, max_row=row+12)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4

    # Colores de las series
    chart1.series[0].graphicalProperties.solidFill = GREEN_PRIMARY
    chart1.series[1].graphicalProperties.solidFill = ORANGE_ACCENT

    ws.add_chart(chart1, "A23")

    # Gráfico de línea - Utilidad mensual
    chart2 = LineChart()
    chart2.title = "Evolución de Utilidad"
    chart2.y_axis.title = "Soles (S/)"
    chart2.style = 10
    chart2.width = 20
    chart2.height = 12

    data2 = Reference(ws, min_col=4, min_row=row, max_col=4, max_row=row+12)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.series[0].graphicalProperties.line.solidFill = BLUE_INFO
    chart2.series[0].graphicalProperties.line.width = 28000

    ws.add_chart(chart2, "A40")

    # Gráfico de torta - Distribución de gastos
    ws.cell(row=57, column=1, value="CATEGORÍA").font = header_font
    ws.cell(row=57, column=1).fill = header_fill
    ws.cell(row=57, column=2, value="MONTO").font = header_font
    ws.cell(row=57, column=2).fill = header_fill

    cats_gasto = ["Personal", "Alquiler", "Insumos", "Servicios", "Marketing", "Otros"]
    montos = [65000, 36000, 28000, 22000, 15000, 12420]
    for i, (cat, monto) in enumerate(zip(cats_gasto, montos)):
        ws.cell(row=58+i, column=1, value=cat)
        ws.cell(row=58+i, column=2, value=monto)
        ws.cell(row=58+i, column=2).number_format = '#,##0'

    pie = PieChart()
    pie.title = "Distribución de Gastos"
    pie.style = 10
    pie.width = 16
    pie.height = 12

    data3 = Reference(ws, min_col=2, min_row=57, max_row=63)
    cats3 = Reference(ws, min_col=1, min_row=58, max_row=63)
    pie.add_data(data3, titles_from_data=True)
    pie.set_categories(cats3)

    # Colores del pie
    colors_pie = [GREEN_PRIMARY, ORANGE_ACCENT, BLUE_INFO, PURPLE, YELLOW_ACCENT, GRAY_TEXT]
    for i, color in enumerate(colors_pie):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        pie.series[0].data_points.append(pt)

    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = True

    ws.add_chart(pie, "D57")

    # Formato condicional - Margen
    ws.conditional_formatting.add(f"E9:E20",
        CellIsRule(operator='greaterThan', formula=['0.3'],
                   fill=PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')))
    ws.conditional_formatting.add(f"E9:E20",
        CellIsRule(operator='lessThan', formula=['0.2'],
                   fill=PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')))

    # Hoja de Resumen Tributario
    ws2 = wb.create_sheet("Tributario")
    ws2.sheet_properties.tabColor = ORANGE_ACCENT

    add_title_banner(ws2, "🏛️ RESUMEN TRIBUTARIO SUNAT", "Cálculos automáticos según normativa vigente")

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 18

    trib_data = [
        ("CONCEPTO", "BASE", "MONTO"),
        ("Ventas Netas", "=Dashboard!B21", ""),
        ("IGV 18%", "=Dashboard!B21", "=Dashboard!F21"),
        ("Renta 1.5% (Mensual)", "=Dashboard!B21", "=Dashboard!G21"),
        ("EsSalud 9%", "", ""),
        ("ONP 13%", "", ""),
        ("Total Obligaciones", "", "=SUM(C3:C6)"),
    ]

    for i, (concepto, base, monto) in enumerate(trib_data):
        r = 4 + i
        ws2.cell(row=r, column=1, value=concepto)
        ws2.cell(row=r, column=2, value=base)
        ws2.cell(row=r, column=3, value=monto)
        if i == 0:
            style_header_row(ws2, r, 3)
        else:
            for col in range(1, 4):
                ws2.cell(row=r, column=col).border = thin_border
                ws2.cell(row=r, column=col).alignment = center_align
                if col in [2, 3]:
                    ws2.cell(row=r, column=col).number_format = '#,##0'

    # Guardar
    output_dir = "plantillas_demo"
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, "01-Dashboard-Financiero-Bodega.xlsx")
    wb.save(filepath)
    print(f"✅ Creado: {filepath}")
    return filepath


# ══════════════════════════════════════════════════════════════
# 2. CONTROL DE VENTAS
# ══════════════════════════════════════════════════════════════
def crear_control_ventas():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"
    ws.sheet_properties.tabColor = BLUE_INFO

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 15

    add_title_banner(ws, "💰 CONTROL DE VENTAS DIARIO", "Registro completo con IGV y análisis de productos")

    # KPIs
    row = 4
    add_kpi_card(ws, row, 1, "VENTAS HOY", 4850, GREEN_PRIMARY)
    add_kpi_card(ws, row, 3, "VENTAS MES", 128750, BLUE_INFO)
    add_kpi_card(ws, row, 5, "PRODUCTOS VENDIDOS", 342, PURPLE)
    add_kpi_card(ws, row, 7, "TICKET PROMEDIO", "S/ 37.50", YELLOW_ACCENT)

    # Tabla de ventas
    row = 8
    headers = ["FECHA", "PRODUCTO", "CANTIDAD", "P. UNITARIO", "SUBTOTAL", "IGV 18%", "TOTAL", "CLIENTE", "ESTADO"]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i+1, value=h)
    style_header_row(ws, row, 9)

    # Datos de ejemplo
    ventas = [
        ("01/01/2025", "Arroz 5kg", 10, 22.50, None, None, None, "Bodega Don Pepe", "Pagado"),
        ("01/01/2025", "Aceite 1L", 15, 8.90, None, None, None, "Restaurante María", "Pagado"),
        ("01/01/2025", "Azúcar 1kg", 20, 4.50, None, None, None, "Pollería Carlitos", "Pendiente"),
        ("02/01/2025", "Fideos 800g", 25, 3.80, None, None, None, "Bodega Don Pepe", "Pagado"),
        ("02/01/2025", "Atún lata", 30, 6.20, None, None, None, "Farmacia Salud", "Pagado"),
        ("02/01/2025", "Leche evaporada", 18, 5.40, None, None, None, "Taller Mecánico", "Pendiente"),
        ("03/01/2025", "Gaseosa 1.5L", 40, 4.20, None, None, None, "Restaurante María", "Pagado"),
        ("03/01/2025", "Pan ciabatta", 50, 2.80, None, None, None, "Hotel Plaza", "Pagado"),
        ("03/01/2025", "Café 250g", 12, 12.50, None, None, None, "Oficina Central", "Pagado"),
        ("04/01/2025", "Galletas pack", 35, 3.20, None, None, None, "Bodega Don Pepe", "Pagado"),
    ]

    for i, (fecha, prod, cant, precio, _, _, _, cliente, estado) in enumerate(ventas):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=fecha)
        ws.cell(row=r, column=2, value=prod)
        ws.cell(row=r, column=3, value=cant)
        ws.cell(row=r, column=4, value=precio)
        ws.cell(row=r, column=5).value = f"=C{r}*D{r}"
        ws.cell(row=r, column=6).value = f"=E{r}*0.18"
        ws.cell(row=r, column=7).value = f"=E{r}+F{r}"
        ws.cell(row=r, column=8, value=cliente)
        ws.cell(row=r, column=9, value=estado)

        for col in [4, 5, 6, 7]:
            ws.cell(row=r, column=col).number_format = '#,##0.00'

    style_data_area(ws, row+1, row+10, 9)

    # Formato condicional - Estado
    ws.conditional_formatting.add(f"I9:I18",
        CellIsRule(operator='equal', formula=['"Pagado"'],
                   fill=PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')))
    ws.conditional_formatting.add(f"I9:I18",
        CellIsRule(operator='equal', formula=['"Pendiente"'],
                   fill=PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')))

    # Validación de datos - Estado
    dv = DataValidation(type="list", formula1='"Pagado,Pendiente,Anulado"', allow_blank=True)
    dv.error = "Seleccione un estado válido"
    dv.errorTitle = "Estado inválido"
    ws.add_data_validation(dv)
    dv.add(f"I9:I100")

    # Totales
    total_row = row + 11
    ws.cell(row=total_row, column=1, value="TOTALES").font = Font(bold=True, size=12)
    for col in [3, 5, 6, 7]:
        ws.cell(row=total_row, column=col).value = f"=SUM({get_column_letter(col)}{row+1}:{get_column_letter(col)}{row+10})"
        ws.cell(row=total_row, column=col).number_format = '#,##0.00'
        ws.cell(row=total_row, column=col).font = Font(bold=True, size=11)

    # Gráfico de barras - Ventas por producto
    chart = BarChart()
    chart.type = "col"
    chart.title = "Ventas por Producto"
    chart.y_axis.title = "Soles (S/)"
    chart.style = 10
    chart.width = 18
    chart.height = 12

    data = Reference(ws, min_col=7, min_row=row, max_row=row+10)
    cats = Reference(ws, min_col=2, min_row=row+1, max_row=row+10)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = BLUE_INFO

    ws.add_chart(chart, "A22")

    # Gráfico de torta - Por cliente
    pie = PieChart()
    pie.title = "Ventas por Cliente"
    pie.style = 10
    pie.width = 16
    pie.height = 12

    # Datos para pie (resumidos)
    ws.cell(row=38, column=1, value="CLIENTE").font = header_font
    ws.cell(row=38, column=1).fill = header_fill
    ws.cell(row=38, column=2, value="TOTAL").font = header_font
    ws.cell(row=38, column=2).fill = header_fill

    clientes = [("Bodega Don Pepe", 850), ("Restaurante María", 620), ("Pollería Carlitos", 450),
                ("Farmacia Salud", 380), ("Taller Mecánico", 290), ("Hotel Plaza", 280), ("Otros", 980)]
    for i, (cli, total) in enumerate(clientes):
        ws.cell(row=39+i, column=1, value=cli)
        ws.cell(row=39+i, column=2, value=total)

    data_pie = Reference(ws, min_col=2, min_row=38, max_row=45)
    cats_pie = Reference(ws, min_col=1, min_row=39, max_row=45)
    pie.add_data(data_pie, titles_from_data=True)
    pie.set_categories(cats_pie)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True

    ws.add_chart(pie, "E22")

    # Guardar
    filepath = os.path.join("plantillas_demo", "02-Control-Ventas-Bodega.xlsx")
    wb.save(filepath)
    print(f"✅ Creado: {filepath}")
    return filepath


# ══════════════════════════════════════════════════════════════
# 3. PLANILLA DE PERSONAL
# ══════════════════════════════════════════════════════════════
def crear_planilla_personal():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planilla"
    ws.sheet_properties.tabColor = PURPLE

    for col in range(1, 14):
        ws.column_dimensions[get_column_letter(col)].width = 14

    add_title_banner(ws, "👥 PLANILLA DE PERSONAL", "Cálculo automático: AFP, ONP, EsSalud, CTS, Gratificación")

    # KPIs
    row = 4
    add_kpi_card(ws, row, 1, "TOTAL TRABAJADORES", 12, PURPLE)
    add_kpi_card(ws, row, 3, "PLANILLA MENSUAL", 18560, GREEN_PRIMARY)
    add_kpi_card(ws, row, 5, "ESSALUD 9%", 1670, ORANGE_ACCENT)
    add_kpi_card(ws, row, 7, "TOTAL EMPRESA", 20230, BLUE_INFO)

    # Headers
    row = 8
    headers = ["NOMBRE", "DNI", "CARGO", "SUELDO", "AFP/ONP", "ESSALUD 9%", "SUELDO NETO",
               "CTS MENSUAL", "GRATIFICACIÓN", "VACACIONES", "BONO", "TOTAL BRUTO", "DEDUCCIONES"]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i+1, value=h)
    style_header_row(ws, row, 13)

    # Datos de empleados
    empleados = [
        ("Juan Pérez García", "12345678", "Gerente", 3500, "AFP", "Prima"),
        ("María López Silva", "23456789", "Vendedora", 1800, "ONP", ""),
        ("Carlos Ruiz Torres", "34567899", "Almacenero", 1500, "AFP", "Integra"),
        ("Ana Martínez Díaz", "45678901", "Cajera", 1500, "ONP", ""),
        ("Pedro Sánchez Ruiz", "56789012", "Vendedor", 1800, "AFP", "Prima"),
        ("Laura García Pérez", "67890123", "Contadora", 2800, "AFP", "Profuturo"),
        ("Roberto Díaz Luna", "78901234", "Motorizado", 1300, "ONP", ""),
        ("Carmen Silva Torres", "89012345", "Vendedora", 1600, "AFP", "Integra"),
        ("Miguel Ángel Rojas", "90123456", "Seguridad", 1200, "ONP", ""),
        ("Lucía Fernández", "11223344", "Limpieza", 1025, "ONP", ""),
        ("Jorge Morales", "22334455", "Auxiliar", 1100, "AFP", "Prima"),
        ("Patricia Vargas", "33445566", "Vendedora", 1500, "ONP", ""),
    ]

    for i, (nombre, dni, cargo, sueldo, tipo, afp_name) in enumerate(empleados):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=nombre)
        ws.cell(row=r, column=2, value=dni)
        ws.cell(row=r, column=3, value=cargo)
        ws.cell(row=r, column=4, value=sueldo)
        ws.cell(row=r, column=4).number_format = '#,##0'

        # AFP/ONP (13% ONP o 10% AFP + comisión)
        if tipo == "ONP":
            ws.cell(row=r, column=5).value = f"=D{r}*0.13"
        else:
            ws.cell(row=r, column=5).value = f"=D{r}*0.10"  # AFP base

        # EsSalud 9% (paga la empresa)
        ws.cell(row=r, column=6).value = f"=D{r}*0.09"

        # Sueldo neto
        ws.cell(row=r, column=7).value = f"=D{r}-E{r}"

        # CTS (sueldo/12)
        ws.cell(row=r, column=8).value = f"=D{r}/12"

        # Gratificación (sueldo/6)
        ws.cell(row=r, column=9).value = f"=D{r}/6"

        # Vacaciones (sueldo/12)
        ws.cell(row=r, column=10).value = f"=D{r}/12"

        # Bono (ejemplo)
        ws.cell(row=r, column=11, value=0)

        # Total bruto
        ws.cell(row=r, column=12).value = f"=D{r}+F{r}+H{r}+I{r}+J{r}+K{r}"

        # Deducciones
        ws.cell(row=r, column=13).value = f"=E{r}"

        # Formato
        for col in [5, 6, 7, 8, 9, 10, 11, 12, 13]:
            ws.cell(row=r, column=col).number_format = '#,##0.00'

    style_data_area(ws, row+1, row+12, 13)

    # Totales
    total_row = row + 13
    ws.cell(row=total_row, column=1, value="TOTALES").font = Font(bold=True, size=12)
    for col in [4, 5, 6, 7, 8, 9, 10, 11, 12, 13]:
        ws.cell(row=total_row, column=col).value = f"=SUM({get_column_letter(col)}{row+1}:{get_column_letter(col)}{row+12})"
        ws.cell(row=total_row, column=col).number_format = '#,##0.00'
        ws.cell(row=total_row, column=col).font = Font(bold=True, size=11)

    # Gráfico - Sueldos por cargo
    chart = BarChart()
    chart.type = "col"
    chart.title = "Sueldos por Cargo"
    chart.y_axis.title = "Soles (S/)"
    chart.style = 10
    chart.width = 20
    chart.height = 12

    data = Reference(ws, min_col=4, min_row=row, max_row=row+12)
    cats = Reference(ws, min_col=1, min_row=row+1, max_row=row+12)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = PURPLE

    ws.add_chart(chart, "A24")

    # Gráfico - Distribución AFP/ONP
    ws.cell(row=40, column=1, value="RÉGIMEN").font = header_font
    ws.cell(row=40, column=1).fill = header_fill
    ws.cell(row=40, column=2, value="CANTIDAD").font = header_font
    ws.cell(row=40, column=2).fill = header_fill

    ws.cell(row=41, column=1, value="AFP")
    ws.cell(row=41, column=2, value=7)
    ws.cell(row=42, column=1, value="ONP")
    ws.cell(row=42, column=2, value=5)

    pie = PieChart()
    pie.title = "Distribución AFP vs ONP"
    pie.style = 10
    pie.width = 14
    pie.height = 10

    data_pie = Reference(ws, min_col=2, min_row=40, max_row=42)
    cats_pie = Reference(ws, min_col=1, min_row=41, max_row=42)
    pie.add_data(data_pie, titles_from_data=True)
    pie.set_categories(cats_pie)

    pt1 = DataPoint(idx=0)
    pt1.graphicalProperties.solidFill = GREEN_PRIMARY
    pie.series[0].data_points.append(pt1)
    pt2 = DataPoint(idx=1)
    pt2.graphicalProperties.solidFill = ORANGE_ACCENT
    pie.series[0].data_points.append(pt2)

    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True

    ws.add_chart(pie, "E40")

    # Hoja de Boleta de Pago
    ws2 = wb.create_sheet("Boleta de Pago")
    ws2.sheet_properties.tabColor = GREEN_PRIMARY

    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 15

    ws2.merge_cells('A1:B1')
    ws2.cell(row=1, column=1, value="BOLETA DE PAGO").font = Font(bold=True, size=16, color=WHITE)
    ws2.cell(row=1, column=1).fill = PatternFill(start_color=GREEN_DARK, end_color=GREEN_DARK, fill_type='solid')
    ws2.cell(row=1, column=1).alignment = center_align
    ws2.row_dimensions[1].height = 40

    boleta = [
        ("EMPLEADO:", "Juan Pérez García"),
        ("DNI:", "12345678"),
        ("MES:", "Enero 2025"),
        ("", ""),
        ("INGRESOS", ""),
        ("Sueldo Básico", 3500),
        ("Asignación Familiar", 0),
        ("Bonificación", 0),
        ("Total Ingresos", "=B6+B7+B8"),
        ("", ""),
        ("DEDUCCIONES", ""),
        ("AFP (10%)", "=B6*0.10"),
        ("EsSalud (9%)", "=B6*0.09"),
        ("Total Deducciones", "=B12+B13"),
        ("", ""),
        ("NETO A PAGAR", "=B9-B14"),
    ]

    for i, (concepto, valor) in enumerate(boleta):
        r = 3 + i
        ws2.cell(row=r, column=1, value=concepto)
        ws2.cell(row=r, column=2, value=valor)
        if concepto in ["INGRESOS", "DEDUCCIONES"]:
            ws2.cell(row=r, column=1).font = Font(bold=True, size=11, color=GREEN_DARK)
        elif concepto in ["Total Ingresos", "Total Deducciones", "NETO A PAGAR"]:
            ws2.cell(row=r, column=1).font = Font(bold=True, size=11)
            ws2.cell(row=r, column=2).font = Font(bold=True, size=11)
        if isinstance(valor, (int, float)):
            ws2.cell(row=r, column=2).number_format = '#,##0.00'

    # Guardar
    filepath = os.path.join("plantillas_demo", "03-Planilla-Personal-Bodega.xlsx")
    wb.save(filepath)
    print(f"✅ Creado: {filepath}")
    return filepath


# ══════════════════════════════════════════════════════════════
# 4. CONTROL DE INVENTARIO
# ══════════════════════════════════════════════════════════════
def crear_control_inventario():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"
    ws.sheet_properties.tabColor = ORANGE_ACCENT

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 15

    add_title_banner(ws, "📦 CONTROL DE INVENTARIO", "Stock mínimo, alertas y valorización automática")

    # KPIs
    row = 4
    add_kpi_card(ws, row, 1, "PRODUCTOS TOTAL", 156, BLUE_INFO)
    add_kpi_card(ws, row, 3, "STOCK TOTAL", 8420, GREEN_PRIMARY)
    add_kpi_card(ws, row, 5, "STOCK BAJO", 12, RED_ALERT)
    add_kpi_card(ws, row, 7, "VALOR INVENTARIO", 45820, PURPLE)

    # Headers
    row = 8
    headers = ["CÓDIGO", "PRODUCTO", "CATEGORÍA", "STOCK", "STOCK MÍNIMO", "P. COMPRA", "P. VENTA",
               "VALOR TOTAL", "ESTADO"]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i+1, value=h)
    style_header_row(ws, row, 9)

    # Datos
    productos = [
        ("P001", "Arroz 5kg", "Granos", 150, 50, 18.50, 22.50),
        ("P002", "Aceite 1L", "Aceites", 80, 30, 6.80, 8.90),
        ("P003", "Azúcar 1kg", "Azúcares", 200, 60, 3.20, 4.50),
        ("P004", "Fideos 800g", "Pastas", 120, 40, 2.50, 3.80),
        ("P005", "Atún lata", "Conservas", 45, 50, 4.80, 6.20),
        ("P006", "Leche evaporada", "Lácteos", 90, 35, 3.90, 5.40),
        ("P007", "Gaseosa 1.5L", "Bebidas", 200, 80, 2.80, 4.20),
        ("P008", "Pan ciabatta", "Panadería", 25, 40, 1.50, 2.80),
        ("P009", "Café 250g", "Café", 60, 20, 8.50, 12.50),
        ("P010", "Galletas pack", "Snacks", 180, 50, 1.80, 3.20),
        ("P011", "Detergente 1kg", "Limpieza", 15, 30, 5.20, 7.50),
        ("P012", "Papel higiénico", "Higiene", 8, 25, 3.80, 5.40),
    ]

    for i, (cod, prod, cat, stock, minimo, pcompra, pventa) in enumerate(productos):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=cod)
        ws.cell(row=r, column=2, value=prod)
        ws.cell(row=r, column=3, value=cat)
        ws.cell(row=r, column=4, value=stock)
        ws.cell(row=r, column=5, value=minimo)
        ws.cell(row=r, column=6, value=pcompra)
        ws.cell(row=r, column=7, value=pventa)
        ws.cell(row=r, column=8).value = f"=D{r}*F{r}"
        ws.cell(row=r, column=9).value = f'=IF(D{r}<=E{r},"⚠️ STOCK BAJO","✅ OK")'

        for col in [6, 7, 8]:
            ws.cell(row=r, column=col).number_format = '#,##0.00'

    style_data_area(ws, row+1, row+12, 9)

    # Formato condicional - Stock bajo
    ws.conditional_formatting.add(f"D9:D20",
        CellIsRule(operator='lessThanOrEqual', formula=['E9'],
                   fill=PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid'),
                   font=Font(color='D63031', bold=True)))

    # Data bars en stock
    ws.conditional_formatting.add(f"D9:D20",
        DataBarRule(start_type='min', end_type='max',
                    color=GREEN_PRIMARY))

    # Totales
    total_row = row + 13
    ws.cell(row=total_row, column=1, value="TOTALES").font = Font(bold=True, size=12)
    for col in [4, 8]:
        ws.cell(row=total_row, column=col).value = f"=SUM({get_column_letter(col)}{row+1}:{get_column_letter(col)}{row+12})"
        ws.cell(row=total_row, column=col).number_format = '#,##0.00'
        ws.cell(row=total_row, column=col).font = Font(bold=True, size=11)

    # Gráfico - Stock por producto
    chart = BarChart()
    chart.type = "col"
    chart.title = "Stock por Producto"
    chart.y_axis.title = "Unidades"
    chart.style = 10
    chart.width = 20
    chart.height = 12

    data = Reference(ws, min_col=4, min_row=row, max_row=row+12)
    cats = Reference(ws, min_col=2, min_row=row+1, max_row=row+12)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = ORANGE_ACCENT

    ws.add_chart(chart, "A24")

    # Gráfico - Valor por categoría
    ws.cell(row=40, column=1, value="CATEGORÍA").font = header_font
    ws.cell(row=40, column=1).fill = header_fill
    ws.cell(row=40, column=2, value="VALOR").font = header_font
    ws.cell(row=40, column=2).fill = header_fill

    cats_inv = [("Granos", 2775), ("Aceites", 544), ("Azúcares", 640), ("Pastas", 300),
                ("Conservas", 216), ("Lácteos", 351), ("Bebidas", 560), ("Panadería", 37),
                ("Café", 510), ("Snacks", 324), ("Limpieza", 78), ("Higiene", 30)]
    for i, (cat, val) in enumerate(cats_inv):
        ws.cell(row=41+i, column=1, value=cat)
        ws.cell(row=41+i, column=2, value=val)

    pie = PieChart()
    pie.title = "Valor de Inventario por Categoría"
    pie.style = 10
    pie.width = 16
    pie.height = 12

    data_pie = Reference(ws, min_col=2, min_row=40, max_row=52)
    cats_pie = Reference(ws, min_col=1, min_row=41, max_row=52)
    pie.add_data(data_pie, titles_from_data=True)
    pie.set_categories(cats_pie)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True

    ws.add_chart(pie, "E40")

    # Guardar
    filepath = os.path.join("plantillas_demo", "04-Control-Inventario-Bodega.xlsx")
    wb.save(filepath)
    print(f"✅ Creado: {filepath}")
    return filepath


# ══════════════════════════════════════════════════════════════
# 5. FLUJO DE CAJA
# ══════════════════════════════════════════════════════════════
def crear_flujo_caja():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Flujo de Caja"
    ws.sheet_properties.tabColor = YELLOW_ACCENT

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 16

    add_title_banner(ws, "💵 FLUJO DE CAJA PROYECTADO", "Proyección a 12 meses con análisis de punto de equilibrio")

    # KPIs
    row = 4
    add_kpi_card(ws, row, 1, "SALDO INICIAL", 15000, BLUE_INFO)
    add_kpi_card(ws, row, 3, "INGRESOS PROYECTADOS", 312000, GREEN_PRIMARY)
    add_kpi_card(ws, row, 5, "EGRESOS PROYECTADOS", 248000, ORANGE_ACCENT)
    add_kpi_card(ws, row, 7, "SALDO FINAL", 79000, PURPLE)

    # Headers
    row = 8
    headers = ["MES", "INGRESOS", "EGRESOS", "FLUJO NETO", "SALDO ACUMULADO", "IGV POR PAGAR", "RENTA POR PAGAR", "DISPONIBLE"]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i+1, value=h)
    style_header_row(ws, row, 8)

    # Datos
    meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
             "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    ingresos = [22000, 24500, 28000, 26500, 30000, 28500, 32000, 34000, 31000, 28000, 25500, 22000]
    egresos = [18000, 19500, 21000, 20500, 22000, 21500, 23000, 24500, 22500, 21000, 19500, 18500]

    for i, (mes, ing, egr) in enumerate(zip(meses, ingresos, egresos)):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=mes)
        ws.cell(row=r, column=2, value=ing)
        ws.cell(row=r, column=3, value=egr)
        ws.cell(row=r, column=4).value = f"=B{r}-C{r}"
        if i == 0:
            ws.cell(row=r, column=5).value = f"=D{r}+15000"  # Saldo inicial
        else:
            ws.cell(row=r, column=5).value = f"=E{r-1}+D{r}"
        ws.cell(row=r, column=6).value = f"=B{r}*0.18"
        ws.cell(row=r, column=7).value = f"=B{r}*0.015"
        ws.cell(row=r, column=8).value = f"=E{r}-F{r}-G{r}"

        for col in [2, 3, 4, 5, 6, 7, 8]:
            ws.cell(row=r, column=col).number_format = '#,##0'

    style_data_area(ws, row+1, row+12, 8)

    # Totales
    total_row = row + 13
    ws.cell(row=total_row, column=1, value="TOTALES").font = Font(bold=True, size=12)
    for col in [2, 3, 4, 6, 7]:
        ws.cell(row=total_row, column=col).value = f"=SUM({get_column_letter(col)}{row+1}:{get_column_letter(col)}{row+12})"
        ws.cell(row=total_row, column=col).number_format = '#,##0'
        ws.cell(row=total_row, column=col).font = Font(bold=True, size=11)

    # Formato condicional - Flujo negativo
    ws.conditional_formatting.add(f"D9:D20",
        CellIsRule(operator='lessThan', formula=['0'],
                   fill=PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid'),
                   font=Font(color='D63031', bold=True)))
    ws.conditional_formatting.add(f"D9:D20",
        CellIsRule(operator='greaterThan', formula=['0'],
                   fill=PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid'),
                   font=Font(color='00B894', bold=True)))

    # Gráfico - Flujo neto mensual
    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "Flujo Neto Mensual"
    chart1.y_axis.title = "Soles (S/)"
    chart1.style = 10
    chart1.width = 20
    chart1.height = 12

    data1 = Reference(ws, min_col=4, min_row=row, max_row=row+12)
    cats1 = Reference(ws, min_col=1, min_row=row+1, max_row=row+12)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    chart1.series[0].graphicalProperties.solidFill = YELLOW_ACCENT

    ws.add_chart(chart1, "A24")

    # Gráfico - Saldo acumulado
    chart2 = LineChart()
    chart2.title = "Evolución del Saldo Acumulado"
    chart2.y_axis.title = "Soles (S/)"
    chart2.style = 10
    chart2.width = 20
    chart2.height = 12

    data2 = Reference(ws, min_col=5, min_row=row, max_row=row+12)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats1)
    chart2.series[0].graphicalProperties.line.solidFill = PURPLE
    chart2.series[0].graphicalProperties.line.width = 28000

    ws.add_chart(chart2, "A41")

    # Gráfico - Ingresos vs Egresos
    chart3 = BarChart()
    chart3.type = "col"
    chart3.grouping = "stacked"
    chart3.title = "Ingresos vs Egresos"
    chart3.y_axis.title = "Soles (S/)"
    chart3.style = 10
    chart3.width = 20
    chart3.height = 12

    data3 = Reference(ws, min_col=2, min_row=row, max_col=3, max_row=row+12)
    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(cats1)
    chart3.series[0].graphicalProperties.solidFill = GREEN_PRIMARY
    chart3.series[1].graphicalProperties.solidFill = ORANGE_ACCENT

    ws.add_chart(chart3, "A58")

    # Hoja de Punto de Equilibrio
    ws2 = wb.create_sheet("Punto de Equilibrio")
    ws2.sheet_properties.tabColor = RED_ALERT

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 18

    ws2.merge_cells('A1:C1')
    ws2.cell(row=1, column=1, value="⚖️ PUNTO DE EQUILIBRIO").font = Font(bold=True, size=16, color=WHITE)
    ws2.cell(row=1, column=1).fill = PatternFill(start_color=RED_ALERT, end_color=RED_ALERT, fill_type='solid')
    ws2.cell(row=1, column=1).alignment = center_align
    ws2.row_dimensions[1].height = 40

    pe_data = [
        ("CONCEPTO", "VALOR", "FÓRMULA"),
        ("Costos Fijos Mensuales", 12000, "Alquiler + Servicios + Sueldos"),
        ("Costo Variable Unitario", 15, "Promedio por producto"),
        ("Precio de Venta Unitario", 25, "Promedio por producto"),
        ("Margen de Contribución", "=B4-B3", "Precio - Costo Variable"),
        ("Ratio de Contribución", "=B5/B4", "Margen / Precio"),
        ("PUNTO DE EQUILIBRIO (unidades)", "=B2/B6", "Costos Fijos / Ratio"),
        ("PUNTO DE EQUILIBRIO (soles)", "=B7*B4", "Unidades × Precio"),
    ]

    for i, (concepto, valor, formula) in enumerate(pe_data):
        r = 3 + i
        ws2.cell(row=r, column=1, value=concepto)
        ws2.cell(row=r, column=2, value=valor)
        ws2.cell(row=r, column=3, value=formula)
        if i == 0:
            style_header_row(ws2, r, 3)
        else:
            for col in range(1, 4):
                ws2.cell(row=r, column=col).border = thin_border
                ws2.cell(row=r, column=col).alignment = center_align
        if i in [1, 2, 3, 7]:
            ws2.cell(row=r, column=2).number_format = '#,##0'
        elif i in [5]:
            ws2.cell(row=r, column=2).number_format = '0.0%'

    # Guardar
    filepath = os.path.join("plantillas_demo", "05-Flujo-Caja-Bodega.xlsx")
    wb.save(filepath)
    print(f"✅ Creado: {filepath}")
    return filepath


# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("🚀 Generando 5 plantillas premium...\n")

    archivos = []
    archivos.append(crear_dashboard_financiero())
    archivos.append(crear_control_ventas())
    archivos.append(crear_planilla_personal())
    archivos.append(crear_control_inventario())
    archivos.append(crear_flujo_caja())

    print(f"\n✅ ¡Listo! {len(archivos)} plantillas creadas en /plantillas_demo/")
    print("\n📊 Resumen:")
    for a in archivos:
        print(f"  • {a}")
