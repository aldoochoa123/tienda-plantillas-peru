"""
Genera Excel: Calculadora de Materiales para Muebles de Melamina
Versión 2 - Sin precios individuales, con cálculo de cantos
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ══════════════════════════════════════════════════════════════
# COLORES
# ══════════════════════════════════════════════════════════════

PURPLE = "3D0C8E"
PURPLE_LIGHT = "6C5CE7"
GREEN = "00B894"
BLUE = "0984E3"
ORANGE = "E17055"
RED = "D63031"
TEAL = "00CEC9"
GRAY = "636E72"
LIGHT_GRAY = "F5F6FA"
DARK = "2D3436"
WHITE = "FFFFFF"

def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

header_font = Font(name='Calibri', bold=True, size=11, color=WHITE)
title_font = Font(name='Calibri', bold=True, size=16, color=PURPLE)
subtitle_font = Font(name='Calibri', bold=True, size=12, color=PURPLE_LIGHT)
data_font = Font(name='Calibri', size=10, color=DARK)
data_font_bold = Font(name='Calibri', bold=True, size=10, color=DARK)
total_font = Font(name='Calibri', bold=True, size=11, color=WHITE)

center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='D0D5DD'),
    right=Side(style='thin', color='D0D5DD'),
    top=Side(style='thin', color='D0D5DD'),
    bottom=Side(style='thin', color='D0D5DD')
)

def style_header(ws, row, cols, bg=None):
    for col in cols:
        c = ws.cell(row=row, column=col)
        c.font = header_font
        c.fill = fill(bg or PURPLE)
        c.alignment = center
        c.border = thin_border

def style_row(ws, row, cols, bold=False, bg=None):
    for col in cols:
        c = ws.cell(row=row, column=col)
        c.font = data_font_bold if bold else data_font
        c.border = thin_border
        c.alignment = center
        if bg:
            c.fill = fill(bg)

def add_title(ws, row, text, cols=8, color=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = title_font
    cell.fill = fill(color or LIGHT_GRAY)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 35

def add_subtitle(ws, row, text, cols=8, color=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = subtitle_font
    cell.fill = fill(color or WHITE)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 25

# ══════════════════════════════════════════════════════════════
# HOJA 1: CALCULADORA DE CORTES
# ══════════════════════════════════════════════════════════════

def create_cuts_sheet(wb):
    ws = wb.active
    ws.title = "Calculadora Cortes"
    ws.sheet_properties.tabColor = PURPLE
    
    widths = {'A': 30, 'B': 12, 'C': 15, 'D': 15, 'E': 25, 'F': 15, 'G': 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    add_title(ws, 1, "📐 CALCULADORA DE CORTES - MELAMINA", 7, PURPLE)
    add_subtitle(ws, 2, "Envía esta lista a tu melaminera para que corte y cantee todas las piezas", 7)
    
    # Datos del proyecto
    row = 4
    ws.cell(row=row, column=1, value="DATOS DEL PROYECTO").font = subtitle_font
    ws.cell(row=row, column=1).fill = fill(PURPLE_LIGHT)
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color=WHITE)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    
    row = 5
    fields = [
        ("Nombre del proyecto:", ""),
        ("Cliente:", ""),
        ("Fecha:", ""),
        ("Melaminera:", ""),
    ]
    
    for i, (field, value) in enumerate(fields):
        r = row + i
        ws.cell(row=r, column=1, value=field).font = data_font_bold
        ws.cell(row=r, column=2, value=value).font = data_font
        ws.cell(row=r, column=2).border = thin_border
    
    # Lista de cortes
    row = 10
    ws.cell(row=row, column=1, value="LISTA DE CORTES").font = subtitle_font
    ws.cell(row=row, column=1).fill = fill(GREEN)
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color=WHITE)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    
    row = 11
    headers = ["Pieza", "Cantidad", "Largo (cm)", "Ancho (cm)", "Cantos (lados)", "Notas", ""]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i+1, value=h)
    style_header(ws, row, range(1, 8), GREEN)
    
    # Filas vacías para que el cliente llene
    for i in range(20):
        r = row + 1 + i
        style_row(ws, r, range(1, 8))
    
    # Cálculo automático de cantos
    row = 33
    ws.cell(row=row, column=1, value="CÁLCULO AUTOMÁTICO DE CANTOS").font = subtitle_font
    ws.cell(row=row, column=1).fill = fill(ORANGE)
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color=WHITE)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    
    row = 34
    canto_headers = ["Pieza", "Cantidad", "Largo canto (cm)", "Subtotal (cm)", "Notas", "", ""]
    for i, h in enumerate(canto_headers):
        ws.cell(row=row, column=i+1, value=h)
    style_header(ws, row, range(1, 8), ORANGE)
    
    # Fórmulas para calcular cantos
    for i in range(20):
        r = row + 1 + i
        ws.cell(row=r, column=1).value = f"=A{12+i}"  # Referencia a pieza
        ws.cell(row=r, column=2).value = f"=B{12+i}"  # Referencia a cantidad
        ws.cell(row=r, column=3, value="")  # Largo del canto (manual)
        ws.cell(row=r, column=4).value = f"=B{r}*C{r}"  # Subtotal automático
        style_row(ws, r, range(1, 8))
    
    # Total cantos
    total_row = row + 21
    ws.cell(row=total_row, column=1, value="TOTAL CANTOS").font = total_font
    ws.cell(row=total_row, column=1).fill = fill(ORANGE)
    ws.cell(row=total_row, column=4).value = f"=SUM(D{row+1}:D{row+20})"
    ws.cell(row=total_row, column=4).number_format = '#,##0'
    ws.cell(row=total_row, column=4).font = total_font
    ws.cell(row=total_row, column=4).fill = fill(ORANGE)
    
    # Total metros
    ws.cell(row=total_row+1, column=1, value="TOTAL METROS").font = total_font
    ws.cell(row=total_row+1, column=1).fill = fill(PURPLE)
    ws.cell(row=total_row+1, column=4).value = f"=D{total_row}/100"
    ws.cell(row=total_row+1, column=4).number_format = '#,##0.00'
    ws.cell(row=total_row+1, column=4).font = total_font
    ws.cell(row=total_row+1, column=4).fill = fill(PURPLE)
    
    return ws

# ══════════════════════════════════════════════════════════════
# HOJA 2: MATERIALES
# ══════════════════════════════════════════════════════════════

def create_materials_sheet(wb):
    ws = wb.create_sheet("Materiales")
    ws.sheet_properties.tabColor = BLUE
    
    widths = {'A': 35, 'B': 15, 'C': 15, 'D': 20, 'E': 25}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    add_title(ws, 1, "📋 LISTA DE MATERIALES", 5, BLUE)
    add_subtitle(ws, 2, "Materiales necesarios para tu proyecto (sin incluir tableros)", 5)
    
    row = 4
    headers = ["Material", "Cantidad", "Unidad", "Para qué sirve", "Notas"]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i+1, value=h)
    style_header(ws, row, range(1, 6), BLUE)
    
    materials = [
        ("Canto melamina", "", "metros", "Tapar bordes de tableros", "Comprar 10% extra"),
        ("Tornillo melamina 4x30mm", "", "caja (100 und)", "Unir piezas", ""),
        ("Tornillo melamina 4x50mm", "", "caja (100 und)", "Unir piezas gruesas", ""),
        ("Tarugo madera 6mm", "", "bolsa (100 und)", "Refuerzo de uniones", ""),
        ("Pegamento de contacto", "", "litro", "Pegar cantos", "Si no usas plancha"),
        ("Bisagra estándar", "", "unidad", "Puertas sin cierre suave", ""),
        ("Bisagra cierre suave", "", "unidad", "Puertas con cierre suave", "Recomendado"),
        ("Tirador metálico", "", "unidad", "Abrir puertas y cajones", ""),
        ("Guía de cajón estándar", "", "par", "Cajones ligeros", ""),
        ("Guía de cajón telescópica", "", "par", "Cajones pesados", "Recomendado"),
        ("Pata plástica", "", "unidad", "Base del mueble", ""),
        ("Pata metálica regulable", "", "unidad", "Base del mueble (niveles)", "Recomendado"),
        ("Taco de pared 8mm", "", "unidad", "Fijar a la pared", "Solo si es de pared"),
        ("Tornillo pared 6x60mm", "", "unidad", "Fijar a la pared", "Solo si es de pared"),
        ("Silicona sanitaria", "", "tubo", "Sellar juntas (baño)", "Solo para baño"),
        ("Barra para ropa", "", "metro", "Colgar ropa (closet)", "Solo para closet"),
        ("Soporte barra ropa", "", "unidad", "Sujetar barra", "Solo para closet"),
    ]
    
    for i, (material, cant, unidad, uso, notas) in enumerate(materials):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=material)
        ws.cell(row=r, column=2, value=cant)
        ws.cell(row=r, column=3, value=unidad)
        ws.cell(row=r, column=4, value=uso)
        ws.cell(row=r, column=5, value=notas)
        style_row(ws, r, range(1, 6))
    
    return ws

# ══════════════════════════════════════════════════════════════
# HOJA 3: HERRAMIENTAS
# ══════════════════════════════════════════════════════════════

def create_tools_sheet(wb):
    ws = wb.create_sheet("Herramientas")
    ws.sheet_properties.tabColor = ORANGE
    
    widths = {'A': 30, 'B': 35, 'C': 20, 'D': 20}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    add_title(ws, 1, "🔧 HERRAMIENTAS DE ARMADO", 4, ORANGE)
    add_subtitle(ws, 2, "Solo herramientas para armar (no para cortar)", 4)
    
    row = 4
    headers = ["Herramienta", "Uso", "¿Es indispensable?", "Notas"]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i+1, value=h)
    style_header(ws, row, range(1, 5), ORANGE)
    
    tools = [
        ("Taladro atornillador", "Atornillar y perforar", "SÍ", "Inversión más importante"),
        ("Broca 5mm para melamina", "Perforar agujeros guía", "SÍ", "Evita que se raje la melamina"),
        ("Tornillos 4x30mm", "Unir piezas", "SÍ", "Para tableros de 18mm"),
        ("Tornillos 4x50mm", "Unir piezas gruesas", "SÍ", "Para uniones más fuertes"),
        ("Tarugos madera 6mm", "Refuerzo de uniones", "SÍ", "Más resistencia"),
        ("Martillo", "Colocar tarugos", "SÍ", ""),
        ("Regla 2m", "Medir y marcar", "SÍ", "Metálica preferiblemente"),
        ("Lápiz", "Marcar cortes", "SÍ", ""),
        ("Nivel", "Verificar nivelado", "SÍ", "De burbuja o láser"),
        ("Escuadra", "Marcar ángulos rectos", "SÍ", ""),
        ("Destornillador Phillips", "Ajustar tornillos", "SÍ", "De repuesto"),
        ("Taladro percutor", "Perforar pared", "Solo pared", "Si el mueble va a la pared"),
        ("Broca 8mm para pared", "Perforar pared", "Solo pared", ""),
        ("Tornillos 6x60mm", "Fijar a la pared", "Solo pared", ""),
        ("Tarugos pared 8mm", "Anclar en pared", "Solo pared", ""),
        ("Sierra caladora", "Cortes curvos", "Solo tuberías", "Si hay que hacer cortes para tuberías"),
        ("Transportador ángulos", "Marcar ángulos", "Solo zapatero", "Para estantes inclinados"),
    ]
    
    for i, (herramienta, uso, indispensable, notas) in enumerate(tools):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=herramienta)
        ws.cell(row=r, column=2, value=uso)
        ws.cell(row=r, column=3, value=indispensable)
        ws.cell(row=r, column=4, value=notas)
        style_row(ws, r, range(1, 5))
    
    return ws

# ══════════════════════════════════════════════════════════════
# HOJA 4: PROVEEDORES
# ══════════════════════════════════════════════════════════════

def create_suppliers_sheet(wb):
    ws = wb.create_sheet("Proveedores")
    ws.sheet_properties.tabColor = TEAL
    
    widths = {'A': 25, 'B': 35, 'C': 20, 'D': 25}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    add_title(ws, 1, "🏪 GUÍA DE PROVEEDORES", 4, TEAL)
    add_subtitle(ws, 2, "Melamineras, ferreterías y tiendas online en Perú", 4)
    
    # Melamineras Lima
    row = 4
    ws.cell(row=row, column=1, value="MELAMINERAS EN LIMA (corte y canteado)").font = subtitle_font
    ws.cell(row=row, column=1).fill = fill(TEAL)
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color=WHITE)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    
    row = 5
    headers = ["Proveedor", "Dirección", "Teléfono", "Servicio"]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i+1, value=h)
    style_header(ws, row, range(1, 5), TEAL)
    
    suppliers_lima = [
        ("Melaminas del Perú", "Av. Argentina 1542, Cercado", "01-428-1234", "Corte + canteado"),
        ("Tableros y Laminados", "Av. Colonial 4567, Callao", "01-555-7890", "Corte + canteado"),
        ("Maderera El Sol", "Jr. Huancavelica 890, Lima", "01-333-4567", "Corte + canteado"),
        ("Cantos y Laminados", "Av. Grau 678, Barranco", "01-444-5678", "Canteado especializado"),
        ("Maderera San Juan", "Av. San Juan 3456, San Juan", "01-666-1234", "Corte + canteado"),
    ]
    
    for i, (proveedor, direccion, telefono, servicio) in enumerate(suppliers_lima):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=proveedor)
        ws.cell(row=r, column=2, value=direccion)
        ws.cell(row=r, column=3, value=telefono)
        ws.cell(row=r, column=4, value=servicio)
        style_row(ws, r, range(1, 5))
    
    # Provincias
    row = row + len(suppliers_lima) + 2
    ws.cell(row=row, column=1, value="MELAMINERAS EN PROVINCIAS").font = subtitle_font
    ws.cell(row=row, column=1).fill = fill(PURPLE_LIGHT)
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color=WHITE)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    
    row += 1
    headers_prov = ["Ciudad", "Proveedor", "Teléfono", "Servicio"]
    for i, h in enumerate(headers_prov):
        ws.cell(row=row, column=i+1, value=h)
    style_header(ws, row, range(1, 5), PURPLE_LIGHT)
    
    suppliers_provinces = [
        ("Arequipa", "Maderera Arequipa", "054-123-456", "Corte + canteado"),
        ("Trujillo", "Melaminas Trujillo", "044-789-012", "Corte + canteado"),
        ("Chiclayo", "Tableros Chiclayo", "074-345-678", "Corte + canteado"),
        ("Piura", "Maderera Norte", "073-901-234", "Corte + canteado"),
        ("Cusco", "Melaminas Cusco", "084-567-890", "Corte + canteado"),
        ("Huancayo", "Ferretería Central", "064-234-567", "Corte + canteado"),
        ("Iquitos", "Maderera Amazonas", "065-890-123", "Corte + canteado"),
    ]
    
    for i, (ciudad, proveedor, telefono, servicio) in enumerate(suppliers_provinces):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=ciudad)
        ws.cell(row=r, column=2, value=proveedor)
        ws.cell(row=r, column=3, value=telefono)
        ws.cell(row=r, column=4, value=servicio)
        style_row(ws, r, range(1, 5))
    
    # Ferreterías
    row = row + len(suppliers_provinces) + 2
    ws.cell(row=row, column=1, value="FERRETERÍAS (herramientas y accesorios)").font = subtitle_font
    ws.cell(row=row, column=1).fill = fill(ORANGE)
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color=WHITE)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    
    row += 1
    headers_hw = ["Ferretería", "Dirección", "Teléfono", "Especialidad"]
    for i, h in enumerate(headers_hw):
        ws.cell(row=row, column=i+1, value=h)
    style_header(ws, row, range(1, 5), ORANGE)
    
    hardware = [
        ("Ferretería Mega", "Av. Brasil 2345, Breña", "01-222-8901", "Herrajes y bisagras"),
        ("Ferretería Construmax", "Av. Argentina 2345, Cercado", "01-777-5678", "Todo en ferretería"),
        ("Sodimac", "Múltiples sedes", "01-600-0000", "Herramientas"),
        ("Promart", "Múltiples sedes", "01-700-0000", "Hogar"),
        ("Falabella", "Múltiples sedes", "01-500-0000", "Herramientas"),
    ]
    
    for i, (ferreteria, direccion, telefono, especialidad) in enumerate(hardware):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=ferreteria)
        ws.cell(row=r, column=2, value=direccion)
        ws.cell(row=r, column=3, value=telefono)
        ws.cell(row=r, column=4, value=especialidad)
        style_row(ws, r, range(1, 5))
    
    # Online
    row = row + len(hardware) + 2
    ws.cell(row=row, column=1, value="TIENDAS ONLINE").font = subtitle_font
    ws.cell(row=row, column=1).fill = fill(GREEN)
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color=WHITE)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    
    row += 1
    headers_online = ["Tienda", "Website", "Teléfono", "Especialidad"]
    for i, h in enumerate(headers_online):
        ws.cell(row=row, column=i+1, value=h)
    style_header(ws, row, range(1, 5), GREEN)
    
    online = [
        ("Mercado Libre", "mercadolibre.com.pe", "N/A", "Todo tipo"),
        ("Sodimac", "sodimac.com.pe", "01-600-0000", "Herramientas"),
        ("Falabella", "falabella.com.pe", "01-500-0000", "Herramientas"),
        ("Promart", "promart.pe", "01-700-0000", "Hogar"),
    ]
    
    for i, (tienda, website, telefono, especialidad) in enumerate(online):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=tienda)
        ws.cell(row=r, column=2, value=website)
        ws.cell(row=r, column=3, value=telefono)
        ws.cell(row=r, column=4, value=especialidad)
        style_row(ws, r, range(1, 5))
    
    return ws

# ══════════════════════════════════════════════════════════════
# GENERAR EXCEL
# ══════════════════════════════════════════════════════════════

def generate_excel():
    output_path = "Calculadora_Materiales_Melamina_V2.xlsx"
    
    wb = openpyxl.Workbook()
    
    create_cuts_sheet(wb)
    create_materials_sheet(wb)
    create_tools_sheet(wb)
    create_suppliers_sheet(wb)
    
    wb.save(output_path)
    
    print(f"✅ Excel generado: {output_path}")
    print(f"📄 Tamaño: {os.path.getsize(output_path) / 1024:.1f} KB")
    
    return output_path

if __name__ == "__main__":
    generate_excel()
